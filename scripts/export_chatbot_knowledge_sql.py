#!/usr/bin/env python3
"""
Exporta datos de Excel (productos + tratamientos) a un SQL de carga
para la tabla public.chatbot_knowledge_items.

Uso:
  python scripts/export_chatbot_knowledge_sql.py ^
    --products-file "C:\\ruta\\Productos.xlsx" ^
    --treatments-file "C:\\ruta\\Tratamientos.xlsx" ^
    --output-sql "chatbot_knowledge_import.sql" ^
    --output-json "chatbot_knowledge_preview.json"
"""

from __future__ import annotations

import argparse
import json
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def to_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def parse_price_numeric(value: Any) -> str | None:
    text = to_str(value)
    if not text:
        return None

    # Solo convertimos formatos de precio simple (evitamos listas tipo "A -> 10€, B -> 20€")
    match = re.fullmatch(r"\s*([0-9]{1,4}(?:[.,][0-9]{1,2})?)\s*(?:€|eur)?\s*", text.lower())
    if not match:
        return None

    normalized = match.group(1).replace(".", "").replace(",", ".")
    try:
        return str(Decimal(normalized))
    except InvalidOperation:
        return None


def sql_quote(value: Any) -> str:
    if value is None:
        return "NULL"
    return "'" + str(value).replace("'", "''") + "'"


def extract_products(products_file: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(products_file, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [to_str(cell) for cell in rows[0][:9]]
    entries: list[dict[str, Any]] = []

    for row_number, row in enumerate(rows[1:], start=2):
        product_name = to_str(row[0])
        if not product_name:
            continue

        entries.append(
            {
                "kind": "product",
                "name": product_name,
                "subtitle": to_str(row[1]),
                "category": "Productos",
                "price_text": to_str(row[2]),
                "price_numeric": parse_price_numeric(row[2]),
                "details": to_str(row[3]),
                "benefits": to_str(row[4]),
                "ingredients": to_str(row[5]),
                "presentation": to_str(row[6]),
                "usage": to_str(row[7]),
                "frequency": None,
                "duration": None,
                "observations": to_str(row[8]),
                "source_file": products_file.name,
                "source_sheet": sheet.title,
                "source_row": row_number,
                "metadata": {"headers": headers},
            }
        )

    return entries


def extract_treatments(treatments_file: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(treatments_file, data_only=True, read_only=True)
    entries: list[dict[str, Any]] = []

    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        header_index = None
        headers = None
        for idx, row in enumerate(rows[:10]):
            col1 = (to_str(row[0]) or "").lower()
            col2 = (to_str(row[1]) or "").lower()
            if col1.startswith("tratamiento") and col2.startswith("precio"):
                header_index = idx
                headers = [to_str(cell) for cell in row[:4]]
                break

        if header_index is None:
            continue

        col4_header = (headers[3] or "").lower()

        for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
            name = to_str(row[0])
            if not name or name.lower() == "tratamiento":
                continue

            price_text = to_str(row[1])
            details = to_str(row[2])
            col4_value = to_str(row[3])

            frequency = None
            duration = None
            if "frecuencia" in col4_header:
                frequency = col4_value
            elif "duración" in col4_header or "duracion" in col4_header:
                duration = col4_value
            else:
                frequency = col4_value

            if not any([price_text, details, frequency, duration]):
                continue

            entries.append(
                {
                    "kind": "treatment",
                    "name": name,
                    "subtitle": None,
                    "category": sheet.title,
                    "price_text": price_text,
                    "price_numeric": parse_price_numeric(price_text),
                    "details": details,
                    "benefits": None,
                    "ingredients": None,
                    "presentation": None,
                    "usage": None,
                    "frequency": frequency,
                    "duration": duration,
                    "observations": None,
                    "source_file": treatments_file.name,
                    "source_sheet": sheet.title,
                    "source_row": row_number,
                    "metadata": {"headers": headers},
                }
            )

    return entries


def dedupe_entries(entries: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: set[tuple[Any, ...]] = set()
    unique: list[dict[str, Any]] = []

    for entry in entries:
        dedupe_key = (
            entry["kind"],
            entry["name"],
            entry["category"],
            entry["price_text"] or "",
            entry["details"] or "",
            entry["frequency"] or "",
            entry["duration"] or "",
        )
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        unique.append(entry)

    return unique


def build_sql(entries: list[dict[str, Any]]) -> str:
    sql_lines: list[str] = []
    sql_lines.append("begin;")
    sql_lines.append("truncate table public.chatbot_knowledge_items;")
    sql_lines.append(
        "insert into public.chatbot_knowledge_items "
        "(kind, name, subtitle, category, price_text, price_numeric, details, benefits, ingredients, presentation, usage, frequency, duration, observations, source_file, source_sheet, source_row, metadata) values"
    )

    value_rows: list[str] = []
    for entry in entries:
        value_rows.append(
            "("
            + ", ".join(
                [
                    sql_quote(entry["kind"]),
                    sql_quote(entry["name"]),
                    sql_quote(entry["subtitle"]),
                    sql_quote(entry["category"]),
                    sql_quote(entry["price_text"]),
                    entry["price_numeric"] if entry["price_numeric"] is not None else "NULL",
                    sql_quote(entry["details"]),
                    sql_quote(entry["benefits"]),
                    sql_quote(entry["ingredients"]),
                    sql_quote(entry["presentation"]),
                    sql_quote(entry["usage"]),
                    sql_quote(entry["frequency"]),
                    sql_quote(entry["duration"]),
                    sql_quote(entry["observations"]),
                    sql_quote(entry["source_file"]),
                    sql_quote(entry["source_sheet"]),
                    str(entry["source_row"]),
                    sql_quote(json.dumps(entry["metadata"], ensure_ascii=False)) + "::jsonb",
                ]
            )
            + ")"
        )

    sql_lines.append(",\n".join(value_rows) + ";")
    sql_lines.append("commit;")
    return "\n".join(sql_lines)


def main() -> None:
    parser = argparse.ArgumentParser(description="Exporta Excel a SQL para chatbot_knowledge_items")
    parser.add_argument("--products-file", required=True, type=Path)
    parser.add_argument("--treatments-file", required=True, type=Path)
    parser.add_argument("--output-sql", required=True, type=Path)
    parser.add_argument("--output-json", required=False, type=Path)
    args = parser.parse_args()

    product_entries = extract_products(args.products_file)
    treatment_entries = extract_treatments(args.treatments_file)
    raw_entries = product_entries + treatment_entries
    final_entries = dedupe_entries(raw_entries)

    sql_text = build_sql(final_entries)
    args.output_sql.write_text(sql_text, encoding="utf-8")

    if args.output_json:
        preview = {
            "total_raw": len(raw_entries),
            "total_final": len(final_entries),
            "sample": final_entries[:10],
        }
        args.output_json.write_text(json.dumps(preview, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Export complete. Raw={len(raw_entries)} Final={len(final_entries)}")
    print(f"SQL: {args.output_sql}")
    if args.output_json:
        print(f"Preview JSON: {args.output_json}")


if __name__ == "__main__":
    main()
